import openpyxl

workbook = openpyxl.load_workbook(filename="data1.xlsx")
s1 = workbook.active

s2 = workbook.copy_worksheet(s1)
s2.title = 'Sheet1_advanced'

#correcting the test excel sheet (swapping BitsId and Name)
temp = s2.cell(row = 1, column = 2).value
s2.cell(row = 1, column = 2).value = s2.cell(row = 1, column = 1).value
s2.cell(row = 1, column = 1).value = temp

for i in range(2,24):
    val = str(s2.cell(row = i, column = 1).value)
    #Emails Column
    s2.cell(row = 1, column = 3).value = 'Email_Id'
    em = val[8:12]
    s2.cell(row = i, column=3).value = 'f2022' + str(em) + '@pilani.bits-pilani.ac.in'
    #Branch Column
    s2.cell(row = 1, column = 4).value = 'Branch'
    br_check = val[7]

    br_trim = val[4:6]

    match br_trim:
        case 'AA':
            s2.cell(row = i, column = 4).value = 'ECE'
        case 'AB':
            s2.cell(row = i, column =4).value = 'Manu'
        case 'A1':
            s2.cell(row = i, column = 4).value = 'Chemical'
        case 'A2':
            s2.cell(row = i, column = 4).value = 'Civil'
        case 'A3':
            s2.cell(row = i, column = 4).value = 'EEE'
        case 'A4':
            s2.cell(row = i, column = 4).value = 'Mech'
        case 'A5':
            s2.cell(row = i, column = 4).value = 'Pharma'
        case 'A7':
            s2.cell(row = i, column = 4).value = 'CSE'
        case 'A8':
            s2.cell(row = i, column = 4).value = 'ENI'
        case 'B1':
            s2.cell(row = i, column = 4).value = 'MSc Bio'
        case 'B2':
            s2.cell(row = i, column = 4).value = 'MSc Chem'
        case 'B3':
            s2.cell(row = i, column = 4).value = 'MSc Eco'
        case 'B4':
            s2.cell(row = i, column = 4).value = 'MSc Math'
        case 'B5':
            s2.cell(row = i, column = 4).value = 'MSc Physics'
    
    if br_check != 'S':             #dual degree
        br_trim2 = val[6:8]
        match br_trim2:
            case 'A3':
                s2.cell(row = i, column = 4).value = str(s2.cell(row = i, column = 4).value) + ' + EEE'
            case 'A7':
                s2.cell(row = i, column = 4).value = str(s2.cell(row = i, column = 4).value) + ' + CSE'                

workbook.save("Data1.xlsx")

import json
products = {}

# Using the values_only because you want to return the cells' values
for row in s2.iter_rows(min_row=2, max_row=23, min_col=1,max_col=4,values_only=True):
    bits_id = row[0]
    product = {
        "bits_id": row[0],
        "name": row[1],
        "email": row[2],
        "branch": row[3]
    }
    products[bits_id] = product

# Using json here to be able to format the output for displaying later
myjson = json.dumps(products)
#Putting in a file
with open('data.json', 'w') as f:
    f.write(myjson)